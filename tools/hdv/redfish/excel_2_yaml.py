##############################################################################
# Copyright (c) 2020 China Mobile Co.,Ltd and others.
#
# All rights reserved. This program and the accompanying materials
# are made available under the terms of the Apache License, Version 2.0
# which accompanies this distribution, and is available at
# http://www.apache.org/licenses/LICENSE-2.0
##############################################################################
'''
excel 2 yaml tools
convert excel config to yaml format config: depends.yaml and cases.yaml.
'''
import os
import yaml
from openpyxl.reader.excel import load_workbook
# pylint: disable=E0611
from log_utils import LOGGER


def load_sheet(excel_file, sheet_index, start_col, end_col):
    '''
    load sheet
    '''
    if not os.path.exists(excel_file):
        LOGGER.error("excel file not existing")
        return None
    input_file = load_workbook(excel_file)
    input_ws = input_file[input_file.sheetnames[sheet_index]]
    cell_key = []
    rows_list = []
    for i in range(start_col, end_col):
        cell_key.append(input_ws.cell(row=1, column=i).value)
    row = 2
    while input_ws.cell(row=row, column=1).value:
        cell_value = []
        for i in range(start_col, end_col):
            value = input_ws.cell(row=row, column=i).value
            if isinstance(value, str):
                value = value.strip().replace('\n', '')
            cell_value.append(value)
        cell_dict = dict(zip(cell_key, cell_value))
        row += 1
        rows_list.append(cell_dict)

    LOGGER.info(rows_list)
    return rows_list


def create_yaml(id_dict, yaml_file):
    '''
    create yaml
    '''
    with open(yaml_file, 'w') as y_file:
        yaml.dump(id_dict, y_file, explicit_start=True)


DEPEND_FILE_NAME = "./conf/depends.yaml"
LOGGER.info("create %s ", DEPEND_FILE_NAME)
create_yaml(load_sheet("./conf/cases.xlsx", 1, 1, 5), DEPEND_FILE_NAME)

CASE_FILE_NAME = "./conf/cases.yaml"
create_yaml(load_sheet("./conf/cases.xlsx", 0, 1, 10), CASE_FILE_NAME)
