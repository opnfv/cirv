##############################################################################
# Copyright (c) 2020 China Mobile Co.,Ltd and others.
#
# All rights reserved. This program and the accompanying materials
# are made available under the terms of the Apache License, Version 2.0
# which accompanies this distribution, and is available at
# http://www.apache.org/licenses/LICENSE-2.0
##############################################################################
'''
an implementation of hardware delivery validation based on redfish interface.
'''
import time
import os
import re
from re import DOTALL as DT
import json
import copy
from ast import literal_eval
import yaml
from openpyxl.reader.excel import load_workbook
from http_handler import UrllibHttpHandler, HEADERS
# pylint: disable=E0611
from log_utils import BASE_DIR, LOG_FILE, LOGGER
from errors import ERROR_CODE, WARN_CODE

LOGGER.info(BASE_DIR)

ACCOUNT_INFO = {}
WAIT_INTERVAL = 5


def parse_config(config_yaml):
    """
    parse setting from config.yaml
    :return:
    """
    try:
        if not os.path.exists(config_yaml):
            LOGGER.error(" %s, %s", ERROR_CODE['E400001'], config_yaml)
        with open(config_yaml, 'r') as conf_file:
            config = yaml.load(conf_file.read(), Loader=yaml.FullLoader)
    except FileNotFoundError as fnfe:
        LOGGER.error(fnfe)
        LOGGER.error(u"%s", ERROR_CODE['E400002'])
        return None
    else:
        return config


def get_token(http_handler, url):
    """
    :return: x_auth_token
    """
    retry_num = 3
    x_auth_token = None
    while retry_num:
        retry_num -= 1
        res = http_handler.post(url, ACCOUNT_INFO)
        if res is None:
            LOGGER.error("%s, %s", WARN_CODE['W100001'], url)
            LOGGER.info("wait %s seconds to try again", WAIT_INTERVAL)
            time.sleep(WAIT_INTERVAL)
            continue
        data = res.info()
        if "X-Auth-Token" in data:
            x_auth_token = data.get("X-Auth-Token")
            return x_auth_token
        else:
            time.sleep(WAIT_INTERVAL)
    return None


def get_etag(http_handler, url):
    """
    :return: ETag
    """
    etag = None
    res = http_handler.get(url)
    data = None
    if res is not None:
        data = res.info()
    if data is not None and "ETag" in data:
        etag = data.get("ETag")
    return etag


def parse_data(exp_value, act_value):
    '''
    parse the expected value and actual value:
    @return:   case 1: exp_value and actual value is str or int,
                then return tuple (exp_value,act_value)
               case 2: list,dict type, then return updated exp_value
               ERROR_CODE for unexpected case.
    '''
    if isinstance(exp_value, (str, int)) and isinstance(act_value, (str, int)):
        return (exp_value, act_value)
    if isinstance(exp_value, list):
        if not isinstance(act_value, list):
            return (exp_value, act_value)
        else:
            for exp in enumerate(exp_value, start=0):
                index = exp[0]
                exp_value[index] = parse_data(
                    exp_value[index], act_value[index])

    elif isinstance(exp_value, dict):
        if isinstance(act_value, dict):
            for key, val in exp_value.items():
                if key in act_value:
                    exp_value[key] = parse_data(val, act_value[key])
                else:
                    LOGGER.error("%s,%s", ERROR_CODE['E500001'], key)
        else:
            LOGGER.error("%s,expected: %s , actual: %s",
                         ERROR_CODE['E400005'], exp_value, act_value)
    else:
        LOGGER.error("%s, expected type:%s, actual type %s",
                     ERROR_CODE['E400006'], type(exp_value), type(act_value))
    return exp_value


def compare_data(value, flag):
    '''
    compare value content
    '''
    if isinstance(value, tuple):
        if value[1] is not None or value[1]:
            if value[0] == 'N/A':
                return "Success", flag
            elif isinstance(value[0], (bool, int, str)):
                if value[0] == value[1]:
                    return "Success", flag
                else:
                    flag += 1
                    return "Failure, expect value: " + str(value[0]) + \
                        ", return value: " + str(value[1]), flag
            elif value[1] in value[0] or value[0] == ['N/A']:
                return "Success", flag
            else:
                flag += 1
                return "Failure, expect value: " + str(value[0]) + \
                    ", return value: " + str(value[1]), flag
        else:
            flag += 1
            return "Failure, expect value: " + str(value[0]) + \
                ", return value: " + str(value[1]), flag

    elif isinstance(value, list):
        for elem in enumerate(value, start=0):
            index = elem[0]
            value[index], flag = compare_data(value[index], flag)
    elif isinstance(value, dict):
        for key, val in value.items():
            value[key], flag = compare_data(val, flag)
    else:
        LOGGER.error("%s", ERROR_CODE['E400007'])
        flag += 1
    return value, flag


def get_component_ids_yaml(file):
    '''
    get component ids from yaml file
    '''
    if not os.path.exists(file):
        LOGGER.info("%s, %s", ERROR_CODE['E400001'], file)
        return None
    return yaml.load(open(file, "r"))


def get_component_ids_excel(excel_file):
    '''
    get the component_id settings from the excel sheet2
    the componnet_id is the parent id of the hardware resource of sheet1
    '''
    input_file = load_workbook(excel_file)
    input_ws = input_file[input_file.sheetnames[1]]
    cell_key = []
    id_info_list = []
    for i in range(1, 5):
        cell_key.append(input_ws.cell(row=1, column=i).value)
    row = 2
    while input_ws.cell(row=row, column=1).value:
        cell_value = []
        for i in range(1, 5):

            cell_value.append(input_ws.cell(row=row, column=i).value.
                              encode("utf8").decode("utf8").replace('\n', ''))
        cell_dict = dict(zip(cell_key, cell_value))
        row += 1
        id_info_list.append(cell_dict)
    return id_info_list


def create_real_url(url_value, id_dict, config_file):
    '''
    create the real url
    either a static url, or a replaced url by depended_id
    '''
    url_list = []
    replaced = 0
    regexp = r'[^{]*{(?P<var>[a-zA-Z_]*)}'
    # pattern = re.compile(regexp, re.S)
    pattern = re.compile(regexp, DT)
    LOGGER.info("url_value %s", url_value)
    matches = list(pattern.finditer(url_value))
    for match in matches:
        value = match.groupdict()
        if value['var'] in config_file:
            url_value = url_value.replace('{' + str(value['var']) + '}',
                                          str(config_file[value['var']]))

        elif value['var'] in id_dict:
            replaced = 1
            instance_list = id_dict[value['var']]
            for instance in instance_list:
                sgl_url = url_value.replace('{' + str(value['var']) + '}',
                                            str(instance))
                LOGGER.debug("replaced url value %s", sgl_url)
                url_list.append(sgl_url)
        else:
            replaced = 2
            LOGGER.error("%s for parameter %s",
                         ERROR_CODE['E300002'], value['var'])
    # combine single case with list case together.
    if replaced == 0:
        LOGGER.info("adding static url %s into list", url_value)
        url_list.append(url_value)
    return url_list


def execute_get_url(url, http_handler):
    """
    execute the url
    """
    LOGGER.debug("execute url %s", url)
    rsp = http_handler.get(url)
    if rsp is None:
        LOGGER.error("return None for url %s", url)
        return None
    ret_dict = {}
    ret_dict.update({"return_code": rsp.code})
    return_value = json.loads(rsp.read())
    ret_dict.update({"return_value": return_value})
    LOGGER.info("ret_dict is %s", ret_dict)
    LOGGER.debug("ret_dict type is %s", type(ret_dict))
    return ret_dict


def handle_depend_url(method, url_list, http_handler):
    '''
    run request url in url_list and collect the response as list
    '''
    response_list = []
    if method == 'GET':
        for url_case in url_list:
            response = execute_get_url(url_case, http_handler)
            response_list.append(response)
    elif method == 'POST':
        pass
    elif method == 'PATCH':
        pass
    elif method == 'DELETE':
        pass
    return response_list


def create_obj_id_list(key_flags, response_list):
    '''
    create object id list
    '''
    if response_list is None or response_list.__len__() == 0:
        LOGGER.debug("response list is None")
        return None
    if key_flags is not None:
        key_list = key_flags.split(':')
    end_id_list = []
    for response in response_list:
        if response is None:
            LOGGER.warning("response is None")
            continue
        return_value = response['return_value']
        if len(key_list) == 1 and key_list[0] in return_value:
            for i in return_value[key_list[0]]:
                end_id_list.append(i['@odata.id'])
        elif len(key_list) > 1:
            for elem in enumerate(key_list, start=0):
                index = elem[0]
                if index == len(key_list) - 1:
                    for case in return_value[key_list[index]]:
                        end_id_list.append(case['@odata.id'])
                else:
                    if isinstance(return_value, list):
                        return_value = return_value[0]
                    elif isinstance(return_value, dict):
                        return_value = return_value[key_list[index]]
                    else:
                        LOGGER.warning("%s, %s", WARN_CODE['W100002'],
                                       type(return_value))

        else:
            LOGGER.error("%s %s", ERROR_CODE['E400003'], key_flags)
    return end_id_list


def get_depend_id(config_file, http_handler, depend_ids):
    '''
    @param mode: yaml or excel,default value "excel"
    parse the component id list
    build up the id resource for each component_id
    return: id_dict like {component_id:[obj_list]}
    '''
    id_dict = {}
    for case in depend_ids:
        component_name = case.get('component_id')
        LOGGER.info("parsing component %s", component_name)
        pro_value = case.get('pro_value')
        url_value = case.get('url_value')
        key_flags = case.get('key_flags')
        # url_list = []
        url_list = create_real_url(url_value, id_dict, config_file)
        # response_list = []
        response_list = handle_depend_url(pro_value, url_list, http_handler)
        # end_id_list = []
        end_id_list = create_obj_id_list(key_flags, response_list)
        if end_id_list is None or end_id_list.__len__() == 0:
            LOGGER.error("%s,%s", ERROR_CODE['E300003'], component_name)
            continue
        id_dict.update({component_name: end_id_list})
    LOGGER.debug("id_dict content is %s", id_dict)
    return id_dict


def read_row(input_ws, row, config_file):
    '''
    read a row value
    '''
    pro_value = input_ws.cell(row=row, column=config_file["pro_seq"]).value
    url_value = input_ws.cell(row=row, column=config_file["url_seq"]).value
    req_body_value = input_ws.cell(
        row=row, column=config_file["req_body_seq"]).value
    expect_return_code = \
        input_ws.cell(
            row=row, column=config_file["expect_return_code_seq"]).value
    expect_return_value = \
        input_ws.cell(
            row=row, column=config_file["expect_return_value_seq"]).value
    attr_name = input_ws.cell(row=row, column=config_file["attr_name"]).value

    if req_body_value is not None:
        req_body_value = literal_eval(req_body_value)
    if expect_return_code is not None:
        expect_return_code = int(expect_return_code)
    if expect_return_value is not None:
        expect_return_value = literal_eval(expect_return_value)
    return pro_value, url_value, req_body_value, expect_return_code,\
        expect_return_value, attr_name


def execute_post_url(body, handler, url):
    '''
    execute post url
    '''
    LOGGER.debug("execute url %s", url)
    rsp = handler.post(url, body)
    LOGGER.debug("post response %s", rsp)
    if not isinstance(rsp, dict):
        LOGGER.error("%s,%s, expected type %s",
                     ERROR_CODE["E400004"], type(rsp), dict)
        return None
    return rsp


def execute_patch_url(body, http_handler, url):
    '''
    execute patch url
    '''
    etag = get_etag(http_handler, url)
    LOGGER.info("etag %s", etag)
    rsp = http_handler.patch(url, body, etag)
    LOGGER.debug("patch response %s", rsp)
    LOGGER.debug("type response is %s", type(rsp))
    ret_dict = {}
    if rsp is None:
        LOGGER.error("%s %s", ERROR_CODE['E100001'], url)
        ret_dict.update({"return_code": "N/A"})
        ret_dict.update({"return_value": "Failure"})
        return ret_dict
    ret_dict.update({"return_code": rsp.code})
    return_value = json.loads(rsp.read())
    ret_dict.update({"return_value": return_value})
    return ret_dict


def handle_final_url(method, url_list, req_body=None, http_handler=None):
    '''execute the requested url to get the response
    '''
    response_list = []
    if method == 'GET':
        for url_case in url_list:
            rsp = execute_get_url(url_case, http_handler)
            response_list.append(rsp)
    elif method == 'POST':
        if len(url_list) > 1:
            LOGGER.error(ERROR_CODE['E100002'])
            return None
        url_value = url_list[0]
        rsp = execute_post_url(req_body, http_handler, url_value)
        response_list.append(rsp)
    elif method == 'PATCH':
        for url_case in url_list:
            LOGGER.info(url_case)
            temp = execute_patch_url(req_body, http_handler, url_case)
            if temp is not None:
                response_list.append(temp)
    elif method == 'DELETE':
        pass
    LOGGER.info("response_list %s", response_list)
    return response_list


def check_component_cnt(expect_return_value, res_list, result):
    '''
    #check if the component count meet the required.
    '''
    if expect_return_value.__contains__('count'):
        if expect_return_value['count'] == len(res_list):
            result.update({"count": "Success"})
        else:
            result.update({"count":
                           "Failure, the actual num is " + str(len(res_list))})
    else:
        result.update({"count": "N/A for this case"})
    return result


def parse_test_result(expect_return_value, expect_return_code,
                      actual_result_list, final_result):
    '''
    @param expected_return_value expected value set in input excel
    @param expected_return_code expected return code
    @param actual_result_list: actual result run by each url list checking
    @param final_result: returned final result
        parsing the test final_result by comparing expected_value with
        real test final_result value.
    '''
    return_code_list = []
    return_value_list = []
    flag = 0
    final_result = check_component_cnt(expect_return_value,
                                       actual_result_list, final_result)

    for each_result in actual_result_list:
        temp_result = {}
        if each_result is not None:
            LOGGER.debug("current result is %s,result_list is %s",
                         each_result, actual_result_list)
            return_code = each_result["return_code"]
            return_code_list.append(return_code)
            return_value = each_result["return_value"]
            if return_code == expect_return_code:
                code_result = 'Success'
            else:
                code_result = 'Failure'
            temp_result.update({'return_code': code_result})
        else:
            LOGGER.warning("%s ,set failure", WARN_CODE['W100003'])
            temp_result.update({'return_code': 'Failure'})
            return_value_list.append(temp_result)
            flag += 1
            continue

        # parse the actual result according to the expected value hierachy.
        ex_value = copy.deepcopy(expect_return_value)
        exp_act_pairs = {}
        for key, value in ex_value.items():
            if key in return_value:
                exp_act_pairs[key] = parse_data(value, return_value[key])
            elif key == 'count':
                pass
            else:
                LOGGER.error("%s, %s", ERROR_CODE['E500001'], key)
                exp_act_pairs[key] = \
                    (value, "Can't find key {} in return value".format(key))
        LOGGER.debug("real_result:%s", exp_act_pairs)

        # comparing expected result with real result.
        if exp_act_pairs:
            for key, value in exp_act_pairs.items():
                temp_result[key], flag = compare_data(value, flag)
            return_value_list.append(temp_result)
    return return_value_list, return_code_list, final_result, flag


def write_result_2_excel(config_file, input_ws, row, flag, result):
    '''
    write the result back to excel
    '''
    if not result:
        input_ws.cell(row=row, column=config_file["detail_result"],
                      value=str('N/A'))
    else:
        input_ws.cell(row=row, column=config_file["detail_result"],
                      value=str(result))
    if flag == 0:
        input_ws.cell(row=row, column=config_file["final_result"],
                      value=str("Success"))
    else:
        input_ws.cell(row=row, column=config_file["final_result"],
                      value=str("Failure"))
    return row


def execute_final_url(config_file, depends_id, http_handler,
                      method, url, req_body):
    '''
    execute final url to get the request result
    '''
    url_list = create_real_url(url, depends_id, config_file)
    rsp_list = handle_final_url(method, url_list, req_body, http_handler)
    return rsp_list


def run_test_case_yaml(config_file, case_file, depends_id, http_handler):
    '''run test case from cases.yaml
    '''
    LOGGER.info("############### start perform test case #################")
    cases_result = []
    cases = read_yaml(case_file)
    for case in cases:
        method, url, req_body, expected_code, expected_value, tc_name \
            = case['method'], case['url'], case['request_body'], \
            case['expected_code'], case['expected_result'], case['case_name']

        expected_value = literal_eval(expected_value)
        flag = 0
        final_rst = {}
        rsp_list = execute_final_url(config_file, depends_id,
                                     http_handler, method, url, req_body)
        if rsp_list is not None and len(rsp_list) > 0:
            return_value_list, return_code_list, final_rst, flag = \
                parse_test_result(
                    expected_value, expected_code, rsp_list, final_rst)
            final_rst.update({'info': return_value_list})
            LOGGER.debug("return_code_list:%s", return_code_list)
            case['return_code_seq'] = str(return_code_list)
        else:
            LOGGER.error("%s", ERROR_CODE['E600001'])
            flag += 1
        case['final_rst'] = "Success" if flag == 0 else "Failure"
        case['details_result'] = \
            str(final_rst) if len(final_rst) > 0 else "N/A"
        cases_result.append(case)
        LOGGER.info("writing test final_rst for case %s", tc_name)

    write_result_2_yaml(cases_result)

    LOGGER.info("############### end perform test case ###################")


def read_yaml(file):
    '''read a yaml file
    '''
    if not os.path.exists(file):
        LOGGER.info("%s %s", ERROR_CODE['E400001'], file)
        return None
    return yaml.load(open(file, "r"))


def write_result_2_yaml(result):
    '''
    write test result to new report.yaml
    '''
    LOGGER.info("writing to yaml file")
    yaml.safe_dump(result, open("./conf/report.yaml", "w"),
                   explicit_start=True)


def run_test_case_excel(config_file, case_file, depends_id, http_handler):
    '''
    perform the test case one by one,
    and write test final_result back to the excel.
    '''
    LOGGER.info("############### start perform test case #################")
    input_file = load_workbook(case_file)
    input_ws = input_file[input_file.sheetnames[0]]

    row = 2
    while input_ws.cell(row=row, column=1).value:
        method, url, req_body, expected_code, expected_value, tc_name \
            = read_row(input_ws, row, config_file)

        LOGGER.info("run test case ##%s##", tc_name)
        if tc_name == "configure BMC ip in static, ipv4":
            LOGGER.debug("debug")
        flag = 0
        final_result = {}
        rsp_list = []
        rsp_list = execute_final_url(config_file, depends_id, http_handler,
                                     method, url, req_body)
        if rsp_list is not None and len(rsp_list) > 0:
            return_value_list, return_code_list, final_result, flag = \
                parse_test_result(expected_value, expected_code,
                                  rsp_list, final_result)
            final_result.update({'info': return_value_list})
            LOGGER.debug("return_code_list:%s", return_code_list)
            input_ws.cell(row=row, column=config_file["return_code_seq"],
                          value=str(return_code_list))
        else:
            LOGGER.error("%s", ERROR_CODE['E600001'])
            flag += 1

        LOGGER.info("writing test final_result for row %s", row)
        row = write_result_2_excel(
            config_file, input_ws, row, flag, final_result)
        row += 1
        input_file.save(case_file)
    LOGGER.info("############### end perform test case ###################")


def run(conf_file, case_excel_file=None, depend_yaml_file=None,
        case_yaml_file=None, file_mode=None):
    '''
    @param conf_file: config.yaml
    @param case_excel_file: excel case file
    @param depend_yaml_file: depends yaml file used if file_mode=yaml
    @param case_yaml_file: case yaml file, used if file_mode=yaml
    @param file_mode: "excel" or "yaml"
    access function
    '''
    # parse config.yaml
    LOGGER.info("start engine ...")
    config_file = parse_config(conf_file)
    http_handler = UrllibHttpHandler()

    # get bmc info
    bmc_ip, bmc_user, bmc_pwd = \
        config_file["bmc_ip"], config_file["bmc_user"], config_file["bmc_pwd"]
    ACCOUNT_INFO.update({"UserName": bmc_user})
    ACCOUNT_INFO.update({"Password": bmc_pwd})

    url = "https://{0}/redfish/v1/SessionService/Sessions".format(bmc_ip)
    x_auth_token = get_token(http_handler, url)
    LOGGER.info("x_auth_token: %s", x_auth_token)

    if x_auth_token is None:
        LOGGER.error("%s token is None", ERROR_CODE['E300001'])
        return None

    HEADERS.update({"X-Auth-Token": x_auth_token})
    id_info_list = None
    if file_mode == "excel":
        id_info_list = get_component_ids_excel(case_excel_file)
    elif file_mode == "yaml":
        id_info_list = get_component_ids_yaml(depend_yaml_file)
    else:
        LOGGER.error("%s,%s", ERROR_CODE['E200001'], file_mode)
        return None

    # get dependent id
    depends_id = get_depend_id(config_file, http_handler, id_info_list)

    # read the test case sheet and perform test
    if file_mode == "excel":
        run_test_case_excel(config_file,
                            case_excel_file, depends_id, http_handler)
    elif file_mode == "yaml":
        run_test_case_yaml(config_file,
                           case_yaml_file, depends_id, http_handler)
    else:
        LOGGER.error("%s,%s", ERROR_CODE['E200001'], file_mode)
        return None

    LOGGER.info("done,checking the log %s", LOG_FILE)

    return True
